import numpy as np
from openpyxl import load_workbook
from packages import fctAnnexes


def vent_apparent(Vt, Vs, Lambda, angle_allure, phi):
    """Fonction calculant le vent apparent percu par les voiles à la gîte, à partir du vent réel, du vent vitesse, et de
     l'angle de dérive et d'allure
    --------------------
    --------------------
    ENTREES :
    Vs : vitesse du bateau, en m/s
    Vt : vitesse du vent reel, en m/s
    Lambda : angle de dérive du bateau, en rad
    angle_allure : true wind angle, angle que prend le vent avec l'axe longitudinal x du bateau, en rad
    phi : angle de gite, en rad
    --------------------
    SORTIES :
    Va : Vitesse du vent apparent perçu par les voiles à la gîte, en m/s
    """
    beta_t = Lambda + angle_allure  # beta_t : page 33 de [2] - angle entre axe X d'avance du bateau et vent réel
    Va = ((Vt * np.sin(beta_t - Lambda) * np.cos(phi)) ** 2 + (Vt * np.cos(beta_t - Lambda) + Vs) ** 2) ** 0.5
    # print("VentApparent =",Va)
    return Va


def polaire_voile(Vt, Vs, rho_air, Lambda, angle_allure, phi, workbook):
    """
    Fonction qui renvoie le torseur dû aux forces aérodynamiques et les coordonnées du centre de poussée, dans le repère
    du fluide en écoulement et en l'origine de ce repère (le centre de poussée)
    --------------------
    ENTREES :
    Vs : vitesse du bateau, en m/s
    Vt : vitesse du vent reel, en m/s
    rho_air : masse volumique de l'air, en kg/m3
    Lambda : angle de dérive du bateau, en rad
    delta_v : angle de réglage de la voile, soit l'angle entre l'axe de la voile, et l'axe longitudinal du bateau, en rad
    angle_allure : true wind angle, angle que prend le vent avec l'axe longitudinal x du bateau, en rad
    phi : angle de gite, en rad
    --------------------
    SORTIES :
    torseur_aero_repFluide : torseur des efforts aérodynamiques dans le repère lié à la voile, en son origine
    CP_repVoile : coordonnées dans le repère lié à la voile du centre de poussée
    """
    Va = vent_apparent(Vt, Vs, Lambda, angle_allure, phi)
    Va_nds = fctAnnexes.ms2nds(Va)
    beta_t = Lambda + angle_allure  # beta_t : page 33 de [2] - angle entre axe X d'avance du bateau et vent réel"
    betaMOINSlambda = np.pi / 2 - np.arctan((Vt * np.cos(beta_t - Lambda) + Vs) / (
                Vt * np.sin(beta_t - Lambda) * np.cos(
            phi)))  # beta : page 33 de [2] - angle entre axe X d'avance du bateau et vent apparent"

    # DETERMINATION DU L'INCIDENCE OPTIMALE alpha_opt A LA VITESSE DE VENT APPARENT CALCULEE
    v_bateau = [1, 5, 10, 15, 20, 25, 30, 35, 40]
    if int(Va_nds) in v_bateau:
        sheet = workbook[str(int(Va_nds)) + "kts"]  # On ouvre la feuille de calcul correspondant au vent
        alpha_opt = np.deg2rad(sheet["N2"].value)
    else:
        i = 0
        Va_inf = v_bateau[i]
        Va_sup = v_bateau[i + 1]
        if Va_nds <= v_bateau[-1]:
            while Va_nds > Va_sup:
                i += 1
                Va_inf = v_bateau[i]
                Va_sup = v_bateau[i + 1]
        else:
            Va_inf = v_bateau[-1]
            Va_sup = v_bateau[-2]
        sheet_inf = workbook[str(Va_inf) + "kts"]  # On ouvre la feuille de calcul correspondant a Vainf
        sheet_sup = workbook[str(Va_sup) + "kts"]  # On ouvre la feuille de calcul correspondant a Vasup
        alpha_opt_Vainf = np.deg2rad(sheet_inf["N2"].value)
        alpha_opt_Vasup = np.deg2rad(sheet_sup["N2"].value)
        alpha_opt = fctAnnexes.interpol_lin(Va_nds, Va_inf, Va_sup, alpha_opt_Vainf, alpha_opt_Vasup)

    # DETERMINATION DU L'INCIDENCE alpha ET DE L'ANGLE DE REGLAGE DE VOILE deltav REELS
    deltav_opt = betaMOINSlambda - alpha_opt  # angle de réglage de voile pour avoir l'incidence optimale
    if abs(deltav_opt) <= np.deg2rad(90):
        # La voile est maintenue à incidence optimale par le flap
        deltav = deltav_opt
        alpha = alpha_opt
    elif deltav_opt < np.deg2rad(-90):
        # Au portant, la voile est en butée et on n'est plus à incidence optimale
        deltav = np.deg2rad(-90)
        alpha = betaMOINSlambda + deltav
    else:
        # Au portant, la voile est en butée et on n'est plus à incidence optimale
        deltav = np.deg2rad(90)
        alpha = betaMOINSlambda - deltav
    # print("alpha_opt = ", np.rad2deg(alpha_opt), "--", "deltaopt",np.rad2deg(deltav_opt))

    # RECUPERATION DES COEFFICIENTS DE PORTANCE ET DE TRAINEE
    if Va_nds < 1:  # Moins de 1 noeud de vent apparent.
        CL = 0
        CD = 0
        S = 0
        CP_repVoile = np.array([0, 0, 0]).reshape(-1, 1)  # le centre de poussée n'a aucune importance
    else:  # Vent apparent supérieur à 1 noeud
        if int(Va_nds) in v_bateau:
            sheet = workbook[str(int(Va_nds)) + "kts"]  # On ouvre la feuille de calcul correspondant au vent
            Lignes = fctAnnexes.trouver_lignes(workbook, alpha, Va_nds)
            n1, n2 = Lignes[0], Lignes[1]
            # Extraction données du fichier excel
            alpha_inf = sheet["B" + str(n1)].value
            alpha_sup = sheet["B" + str(n2)].value
            CL_inf = sheet["F" + str(n1)].value
            CL_sup = sheet["F" + str(n2)].value
            CD_inf = sheet["H" + str(n1)].value
            CD_sup = sheet["H" + str(n2)].value
            XCP_inf = sheet["L" + str(n1)].value
            XCP_sup = sheet["L" + str(n2)].value
            ZCP_inf = sheet["K" + str(n1)].value
            ZCP_sup = sheet["K" + str(n2)].value

            # Interpolation avec les données extraites
            CL = fctAnnexes.interpol_lin(alpha, alpha_inf, alpha_sup, CL_inf, CL_sup)
            CD = fctAnnexes.interpol_lin(alpha, alpha_inf, alpha_sup, CD_inf, CD_sup)
            XCP = fctAnnexes.interpol_lin(alpha, alpha_inf, alpha_sup, XCP_inf, XCP_sup)
            ZCP = fctAnnexes.interpol_lin(alpha, alpha_inf, alpha_sup, ZCP_inf, ZCP_sup)
            CP_repVoile = np.array([XCP, 0, ZCP]).reshape(-1, 1)

        else:
            i = 0
            Va_inf = v_bateau[i]
            Va_sup = v_bateau[i + 1]
            # Limite ajoutée correspond à la vitesse max des données du tableur, après ces limites les données sont
            # interpolées linéairement
            if Va_nds <= v_bateau[-1]:
                while Va_nds > Va_sup:
                    i += 1
                    Va_inf = v_bateau[i]
                    Va_sup = v_bateau[i + 1]
            else:
                Va_inf = v_bateau[-1]
                Va_sup = v_bateau[-2]
            # Va_inf et Va_sup encadrent la valeur de vent apparent, par des valeurs où la polaire de voile est connue.
            Lignes1, Lignes2 = fctAnnexes.trouver_lignes(workbook, alpha, [Va_inf, Va_sup])
            sheet_inf = workbook[str(Va_inf) + "kts"]  # On ouvre la feuille de calcul correspondant a Vainf
            sheet_sup = workbook[str(Va_sup) + "kts"]  # On ouvre la feuille de calcul correspondant a Vasup
            n1, n2 = Lignes1[0], Lignes1[1]  # Pour la feuille de calcul Va_inf
            m1, m2 = Lignes2[0], Lignes2[1]  # Pour la feuille de calcul Va_sup
            # print(n1, n2, m1, m2)
            # Extraction données du fichier excel
            alpha_inf_Vainf = sheet_inf["B" + str(n1)].value
            alpha_sup_Vainf = sheet_inf["B" + str(n2)].value
            alpha_inf_Vasup = sheet_sup["B" + str(m1)].value
            alpha_sup_Vasup = sheet_sup["B" + str(m2)].value
            CL_inf_Vainf = sheet_inf["F" + str(n1)].value
            CL_sup_Vainf = sheet_inf["F" + str(n2)].value
            CL_inf_Vasup = sheet_sup["F" + str(m1)].value
            CL_sup_Vasup = sheet_sup["F" + str(m2)].value
            CD_inf_Vainf = sheet_inf["H" + str(n1)].value
            CD_sup_Vainf = sheet_inf["H" + str(n2)].value
            CD_inf_Vasup = sheet_sup["H" + str(m1)].value
            CD_sup_Vasup = sheet_sup["H" + str(m2)].value
            XCP_inf_Vainf = sheet_inf["L" + str(n1)].value
            XCP_sup_Vainf = sheet_inf["L" + str(n2)].value
            XCP_inf_Vasup = sheet_sup["L" + str(m1)].value
            XCP_sup_Vasup = sheet_sup["L" + str(m2)].value
            ZCP_inf_Vainf = sheet_inf["K" + str(n1)].value
            ZCP_sup_Vainf = sheet_inf["K" + str(n2)].value
            ZCP_inf_Vasup = sheet_sup["K" + str(m1)].value
            ZCP_sup_Vasup = sheet_sup["K" + str(m2)].value
            # NB : la surface est prise que pour une des valeurs de vent, car elle est la même quelle que soit le Va.

            # Première interpolation selon les angles, sur chaque feuille de calcul séparément
            CL_Vainf = fctAnnexes.interpol_lin(alpha, alpha_inf_Vainf, alpha_sup_Vainf, CL_inf_Vainf, CL_sup_Vainf)
            CL_Vasup = fctAnnexes.interpol_lin(alpha, alpha_inf_Vasup, alpha_sup_Vasup, CL_inf_Vasup, CL_sup_Vasup)
            CD_Vainf = fctAnnexes.interpol_lin(alpha, alpha_inf_Vainf, alpha_sup_Vainf, CD_inf_Vainf, CD_sup_Vainf)
            CD_Vasup = fctAnnexes.interpol_lin(alpha, alpha_inf_Vasup, alpha_sup_Vasup, CD_inf_Vasup, CD_sup_Vasup)
            XCP_Vainf = fctAnnexes.interpol_lin(alpha, alpha_inf_Vainf, alpha_sup_Vainf, XCP_inf_Vainf, XCP_sup_Vainf)
            XCP_Vasup = fctAnnexes.interpol_lin(alpha, alpha_inf_Vasup, alpha_sup_Vasup, XCP_inf_Vasup, XCP_sup_Vasup)
            ZCP_Vainf = fctAnnexes.interpol_lin(alpha, alpha_inf_Vainf, alpha_sup_Vainf, ZCP_inf_Vainf, ZCP_sup_Vainf)
            ZCP_Vasup = fctAnnexes.interpol_lin(alpha, alpha_inf_Vasup, alpha_sup_Vasup, ZCP_inf_Vasup, ZCP_sup_Vasup)
            # Seconde interpolation, selon les vitesses
            CL = fctAnnexes.interpol_lin(Va_nds, Va_inf, Va_sup, CL_Vainf, CL_Vasup)
            CD = fctAnnexes.interpol_lin(Va_nds, Va_inf, Va_sup, CD_Vainf, CD_Vasup)
            XCP = fctAnnexes.interpol_lin(Va_nds, Va_inf, Va_sup, XCP_Vainf, XCP_Vasup)
            ZCP = fctAnnexes.interpol_lin(Va_nds, Va_inf, Va_sup, ZCP_Vainf, ZCP_Vasup)
            CP_repVoile = np.array([XCP, 0, ZCP]).reshape(-1, 1)
    S = 1.647
    L = 0.5 * rho_air * CL * S * (Va ** 2)

    # print(alpha)
    # print("L=",L)
    # if L<0:
    #    raise ValueError("Lift is negative !")
    D = 0.5 * rho_air * CD * S * (Va ** 2)
    torseur_aero_repFluide = np.concatenate((np.array([D, L, 0]).reshape(-1, 1), np.array([0, 0, 0]).reshape(-1, 1)),
                                            axis=1)
    return torseur_aero_repFluide, CP_repVoile, deltav


def chgt_base_pt_application(CP_repVoile, O_repVoile_repBateau, deltav, Lambda, phi):
    """
    Fonction qui passe les forces véliques du repère lié au vent à celui lié à l'avance du bateau.
    --------------------
    ENTREES :
    CP_repVoile : coordonnées du point d'application des forces véliques dans le repère lié à la voile
                  np.array de taille (3,1) au format [X_repVoile, Y_repVoile, Z_repVoile].T
    O_repVoile_repBateau : Vecteur donnant les coordonnées de l'origine du repère lié à la voile, dans le repère lié au bateau.
                           vecteur du même format que CP_repVoile
    deltav : Angle d'ouverture de la voile, prix comme l'angle entre -y (repère lié au bateau), et l'axe de la voile
    Lambda : angle de dérive, angle entre axe x longitudinal du bateau et axe X d'avance du bateau
    phi : angle de gîte, angle entre la verticale terrestre et l'axe Z du repère d'avance du bateau
    --------------------
    SORTIES :
    CP_repAvBateau  : coordonnées du point d'application des forces véliques dans le repère d'avance du bateau.
                      vecteur du même format que CP_repVoile.
    """
    ###############################################################
    # ETAPE 1 : Passage du repère lié au vent à celui lié au bateau.
    # Définition de la matrice de passage de repVoile à repBateau
    M_repVoile_repBateau = np.zeros((3, 3))
    M_repVoile_repBateau[0][0] = -np.cos(deltav)
    M_repVoile_repBateau[0][1] = np.sin(deltav)
    M_repVoile_repBateau[1][0] = -np.sin(deltav)
    M_repVoile_repBateau[1][1] = -np.cos(deltav)
    M_repVoile_repBateau[2][2] = 1
    # Calcul des nouvelles coordonnées de CP
    CP_repBateau = M_repVoile_repBateau @ CP_repVoile + O_repVoile_repBateau
    ###############################################################
    # ETAPE 2 : Passage du repère lié au bateau à celui lié à l'avance du bateau
    # Définition de la matrice de passage de repBateau à repAvBateau
    M_repBateau_repAvBateau = np.zeros((3, 3))
    cL, sL, cP, sP = np.cos(Lambda), np.sin(Lambda), np.cos(phi), np.sin(phi)
    M_repBateau_repAvBateau[0][0] = cL
    M_repBateau_repAvBateau[0][1] = -cP * sL
    M_repBateau_repAvBateau[0][2] = sP * sL
    M_repBateau_repAvBateau[1][0] = sL
    M_repBateau_repAvBateau[1][1] = cL * cP
    M_repBateau_repAvBateau[1][2] = -sP * cL
    M_repBateau_repAvBateau[2][1] = sP
    M_repBateau_repAvBateau[2][2] = cP
    # Calcul des nouvelles coordonnées de CP
    CP_repAvBateau = M_repBateau_repAvBateau @ CP_repBateau
    return CP_repAvBateau


def chgt_base_forces(Vt, Vs, Lambda, angle_allure, phi, torseur_aero_repFluide):
    """
    Fonction qui passe les forces véliques du repère lié au vent à celui lié à l'avance du bateau.
    --------------------
    ENTREES :
    Vt : vent réel, en m/s
    Vs : vent vitesse ( = vitesse du bateau), en m/s
    Lambda : angle de dérive, angle entre axe x longitudinal du bateau et axe X d'avance du bateau, en rad
    angle_allure : angle entre axe x longitudinal du bateau et vent réel, en rad
    phi : angle de gîte, angle entre la verticale terrestre et l'axe Z du repère d'avance du bateau, en rad
    torseur_aero_repFluide : torseur aérodynamique donné dans le repère associé à l'écoulement du fluide.
                            Torseur de la forme : [[Fx, Fy, Fz],[Mx, My, Mz]]
    --------------------
    SORTIES :
    F_repAvBateau : forces véliques dans le repère lié à l'avance du bateau. De la forme np.array([Fx,Fy,Fz].T)
    """
    beta_t = -Lambda + angle_allure  # beta_t : page 33 de [2] - angle entre axe X d'avance du bateau et vent réel"
    beta = np.pi / 2 - np.arctan((Vt * np.cos(beta_t) + Vs) / (Vt * np.sin(beta_t) * np.cos(
        phi)))  # beta : page 33 de [2] - angle entre axe X d'avance du bateau et vent apparent"
    Lvoile, Dvoile = np.abs(torseur_aero_repFluide[1][0]), np.abs(torseur_aero_repFluide[0][0])
    FX = Lvoile * np.sin(beta) - Dvoile * np.cos(beta)
    FY = -np.cos(phi) * (Lvoile * np.cos(beta) + Dvoile * np.sin(beta))
    FZ = -np.sin(phi) * (Lvoile * np.cos(beta) + Dvoile * np.sin(beta))
    F_repAvBateau = np.array([FX, FY, FZ]).reshape(-1, 1)
    return F_repAvBateau


def chgt_base_moments(CP_repAvBateau, F_repAvBateau):
    """
    Fonction qui passe les moments associés aux forces véliques, du repère lié au vent à celui lié à l'avance du bateau.
    --------------------
    ENTREES :
    CP_repAvBateau : coordonnées du centre de poussée dans le repère d'avance du bateau, sous la forme d'un vecteur de taille 3.
    F_repAvBateau : Force vélique exprimée dans le repère d'avance du bateau, sous la forme d'un vecteur de taille 3.
    --------------------
    SORTIES :
    M_repAvBateau : Moment en l'origine du repère d'avance du bateau dû aux forces véliques, exprimé sous forme d'un vecteur de taille 3
    """
    M_repAvBateau = np.cross(CP_repAvBateau.T, F_repAvBateau.T)
    return M_repAvBateau.reshape(-1, 1)


def torseur_Faero(Vt, Vs, angle_allure, Lambda, phi, rho_eau, rho_air, workbook):
    """
    Fonction qui renvoie le torseur des Forces véliques exprimé dans le repère d'avance du bateau, en son origine.
    --------------------
    ENTREES :
    fichier : chaîne de caractère donnant le nom du fichier avec les données d'entrée
    --------------------
    SORTIES :
    T : Torseur des forces véliques dans le repère d'avance du bateau, en son origine. Exprimé sous la forme :
        np.array([[FX, FY, FZ], [MX, MY, MZ]])
    """
    O_repVoile_repBateau = np.array([1.3988, 0., 0.34609]).reshape(-1, 1)  # Repères voile et bateau ont la même origine
    torseur_aero_repFluide, CP_repVoile, deltav = polaire_voile(Vt, Vs, rho_air, Lambda, angle_allure, phi, workbook)
    # print(torseur_aero_repFluide)
    CP_repAvBateau = chgt_base_pt_application(CP_repVoile, O_repVoile_repBateau, deltav, Lambda, phi)
    F_repAvBateau = chgt_base_forces(Vt, Vs, Lambda, angle_allure, phi, torseur_aero_repFluide)
    M_repAvBateau = chgt_base_moments(CP_repAvBateau, F_repAvBateau)
    T = np.concatenate((F_repAvBateau, M_repAvBateau), axis=1)
    return T


if __name__ == "__main__":
    import os

    path = os.getcwd()
    pathdir = os.path.dirname(path)
    workbook = load_workbook(filename=os.path.join(pathdir, "tests/data.xlsx"), data_only=True)
    sheet = workbook.active
    Lambda = np.deg2rad(0)
    angle_allure = np.deg2rad(20)
    Vt = fctAnnexes.nds2ms(25)
    Vs = fctAnnexes.nds2ms(0)
    phi = 0
    rho_air = sheet["G3"].value
    rho_eau = sheet["H3"].value
    fichier = os.path.join(pathdir, "data/Polaires_Voile_data4.xlsm")
    workbook2 = load_workbook(filename=fichier, data_only=True)
    print(polaire_voile(Vt, Vs, rho_air, Lambda, angle_allure, phi, workbook2))
    T = torseur_Faero(Vt, Vs, angle_allure, Lambda, phi, rho_eau, rho_air, workbook2)
    # angle_allure_list = np.linspace(10, 180, 170)
    # angle_allure_rad = np.deg2rad(angle_allure_list)
    # deltav_list = []
    # alpha_list = []
    # for i in angle_allure_rad:
    # beta_t = Lambda + i  # beta_t : page 33 de [2] - angle entre axe X d'avance du bateau et vent réel"
    # betaMOINSlambda = np.pi / 2 - np.arctan((Vt * np.cos(beta_t - Lambda) + Vs) / (Vt * np.sin(beta_t - Lambda) * np.cos(phi)))  # beta : page 33 de [2] - angle entre axe X d'avance du bateau et vent apparent"
    # a, b, deltav = polaire_voile(Vt, Vs, rho_air, Lambda, i, phi)
    # alpha_list.append(betaMOINSlambda-deltav)
    # delta.append(deltav)
# plt.plo(angle_allure_list, np.rad2deg(deltav_list), label ='deltav')
# plt.ylabel("delta (deg)")
# plt.xlabel("Angle d'allure (deg)")
# plt.legend()
# plt.plot(angle_allure_list, np.rad2deg(alpha_list), label='alpha')
# plt.ylabel("alpha (deg)")
# plt.xlabel("Angle d'allure (deg)")
# plt.legend()
# plt.show()
