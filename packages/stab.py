import numpy as np
from openpyxl import load_workbook
from packages import fctAnnexes

def efforts_hydrostat(phi, theta, DELTA, workbook):
    """
    Fonction qui renvoie le torseur résultant des efforts hydrostatiques, au point O et dans le repère lié au bateau.
    Les valeurs d'entrée utilisées sont issues du logiciel GHS.
    --------------------
    ENTREES :
    Phi : angle de gîte du bateau, en radians
    Theta : angle d'assiette du bateau, en radians
    DELTA : déplacement du navire, en kg

    --------------------
    SORTIES :
    torseur_stab_repBateau  : Torseur résultant des efforts hydrostatiques, au point O et dans le repère lié au bateau.
                              Type : numpy.ndarray de taille (3,2)
    """
    if abs(theta) > np.deg2rad(15):
        raise ValueError("Theta is over 15 degrees")
    g = 9.81
    sheet_T = workbook["Trans"]
    lignes_T = fctAnnexes.trouver_lignes_stab(workbook, "Trans", np.rad2deg(phi))
    n1_T, n2_T = lignes_T[0], lignes_T[1]
    GZ_T_inf = sheet_T["B" + str(n1_T)].value
    GZ_T_sup = sheet_T["B"+str(n2_T)].value
    phi_inf = np.deg2rad(sheet_T["A" + str(n1_T)].value)
    phi_sup = np.deg2rad(sheet_T["A"+str(n2_T)].value)
    if phi>=0:
        GZ_T = fctAnnexes.interpol_lin(phi, phi_inf, phi_sup, GZ_T_inf, GZ_T_sup)
    else:
        GZ_T = -fctAnnexes.interpol_lin(-phi, phi_inf, phi_sup, GZ_T_inf, GZ_T_sup)
    RM_T = -GZ_T*DELTA*g # signe négatif car si phi>0, le couple est négatif (cf repères)
    sheet_L = workbook["Longi"]
    lignes_L = fctAnnexes.trouver_lignes_stab(workbook, "Longi", np.rad2deg(theta))
    n1_L, n2_L = lignes_L[0], lignes_L[1]
    GZ_L_inf = sheet_L["B"+str(n1_L)].value
    GZ_L_sup = sheet_L["B"+str(n2_L)].value
    theta_inf = np.deg2rad(sheet_L["A" + str(n1_L)].value)
    theta_sup = np.deg2rad(sheet_L["A"+str(n2_L)].value)
    if theta>=0:
        GZ_L = fctAnnexes.interpol_lin(theta, theta_inf, theta_sup, GZ_L_inf, GZ_L_sup)
    else:
        GZ_L = -fctAnnexes.interpol_lin(-theta, theta_inf, theta_sup, GZ_L_inf, GZ_L_sup)
    RM_L = -GZ_L*DELTA*g
    torseur_stab_repBateau = np.concatenate((np.array([0, 0, 0]).reshape(-1, 1), np.array([RM_T, RM_L, 0]).reshape(-1, 1)),
                                           axis=1)
    return torseur_stab_repBateau

def chgt_base_torseur(torseur_stab_repBateau, Lambda):
    """
    Fonction qui change de base le torseur des efforts hydrostatiques, pour le passer dans le repère d'avance du bateau
    --------------------
    ENTREES :
    torseur_stab_repBateau  : Torseur résultant des efforts hydrostatiques, au point O et dans le repère lié au bateau.
                              Type : numpy.ndarray de taille (3,2)
    Lambda : angle de dérive du bateau, en radians

    --------------------
    SORTIES :
    torseur_stab_repAvBateau : Torseur résultant des efforts hydrostatiques, le repère lié à l'avance du bateau.
                              Type : numpy.ndarray de taille (3,2)
    """
    # RM_T_repBateau, RM_L_repBateau = torseur_stab_repBateau[0,1], torseur_stab_repBateau[1,1]
    # RM_T_x_repAvBateau = RM_T_repBateau*np.cos(Lambda)
    # RM_T_y_repAvBateau = RM_T_repBateau*np.sin(Lambda)
    # RM_L_x_repAvBateau = -RM_L_repBateau*np.sin(Lambda)
    # RM_L_y_repAvBateau = RM_L_repBateau*np.cos(Lambda)
    # torseur_stab_repAvBateau = np.array([[0, RM_T_x_repAvBateau+RM_L_x_repAvBateau], [0, RM_L_y_repAvBateau+RM_T_y_repAvBateau], [0, 0]]
    torseur_stab_repAvBateau = torseur_stab_repBateau
    return torseur_stab_repAvBateau

def torseur_Fstab(DELTA, Lambda, phi, theta, workbook):
    """
    Fonction qui à partir des données d'entrée contenues dans fichier, renvoie le torseur dû aux efforts hydrostatiques
    --------------------
    ENTREES :
    fichier : fichier excel contenant les données d'entrée de la simulation
    --------------------
    SORTIES :
    torseur_stab_repAvBateau : Torseur résultant des efforts hydrostatiques, le repère lié à l'avance du bateau.
        Type : numpy.ndarray de taille (3,2)
    """


    torseur_stab_repBateau = efforts_hydrostat(phi, theta, DELTA, workbook)
    torseur_stab_repAvBateau = chgt_base_torseur(torseur_stab_repBateau, Lambda)
    return torseur_stab_repAvBateau


if __name__ == "__main__":
    import os
    path = os.getcwd()
    pathdir = os.path.dirname(path)
    workbook = load_workbook(filename=os.path.join(pathdir,"tests/data.xlsx"), data_only=True)
    sheet = workbook.active
    DELTA = sheet["I3"].value
    Lambda = sheet["C3"].value
    phi = sheet["E3"].value
    theta = sheet["J3"].value
    workbookStab = load_workbook(filename=os.path.join(pathdir,"data/Stab_data.xlsx"), data_only=True)
    T = torseur_Fstab(DELTA, Lambda, phi, theta, workbookStab)
    print(T)
