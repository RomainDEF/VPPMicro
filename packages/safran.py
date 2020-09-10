import numpy as np
from openpyxl import load_workbook
from packages import fctAnnexes

def vitesse_fluide(Vs, Lambda, phi):
    """Fonction calculant le vent vitesse perçu par les appendices à la gîte
    --------------------
    ENTREES :
    Vs : vitesse du bateau, en m/s
    Lambda : angle de dérive du bateau, en rad
    phi : angle de gite, en rad
    --------------------
    SORTIES :
    Vf : Vitesse du fluide percu par les appendices à la gîte, en m/s
    """
    Vf = Vs*(np.cos(Lambda)**2 + np.sin(Lambda)**2 * np.cos(phi)**2)**0.5
    return Vf

def polaire_safran(Vs, rho_eau, Lambda, phi, workbook):
    alpha = np.arctan(np.tan(Lambda) * np.cos(phi))  # alpha = angle d'incidence entre le sqfrqn et le flux d'eau
    Vf = vitesse_fluide(Vs, Lambda, phi)
    Vf_nds = fctAnnexes.ms2nds(Vf)

    if Vf_nds < 0.1 : #Moins de 0.1 noeud de vitesse du bateau
        CL = 0
        CD = 0
        S = 0
        CP_repsafran = np.array([0, 0, 0]).reshape(-1, 1) #le centre de poussée n'a aucune importance
    else : # Vitesse du bateau supérieure à 0.1 noeud
        v_bateau = [0.1, 0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 5, 10]
        if Vf_nds in v_bateau:
            sheet = workbook[str(Vf_nds) + "kts"]  # On ouvre la feuille de calcul correspondant à la vitesse du bateau
            Lignes = fctAnnexes.trouver_lignes(workbook, alpha, Vf_nds)
            n1, n2 = Lignes[0], Lignes[1]
            # Extraction données du fichier excel
            alpha_inf = sheet["B"+str(n1)].value
            alpha_sup = sheet["B"+str(n2)].value
            CL_inf = sheet["F"+str(n1)].value
            CL_sup = sheet["F"+str(n2)].value
            CD_inf = sheet["H" + str(n1)].value
            CD_sup = sheet["H" + str(n2)].value
            XCP_inf = sheet["J"+str(n1)].value
            XCP_sup = sheet["J"+str(n2)].value
            ZCP_inf = sheet["K"+str(n1)].value
            ZCP_sup = sheet["K"+str(n2)].value

            # Interpolation avec les données extraites
            CL = fctAnnexes.interpol_lin(alpha, alpha_inf, alpha_sup, CL_inf, CL_sup)
            CD = fctAnnexes.interpol_lin(alpha, alpha_inf, alpha_sup, CD_inf, CD_sup)
            XCP = fctAnnexes.interpol_lin(alpha, alpha_inf, alpha_sup, XCP_inf, XCP_sup)
            ZCP = fctAnnexes.interpol_lin(alpha, alpha_inf, alpha_sup, ZCP_inf, ZCP_sup)
            CP_repsafran = np.array([XCP, 0, ZCP]).reshape(-1, 1)

        else:
            i = 0
            Vf_inf = v_bateau[i]
            Vf_sup = v_bateau[i+1]
            while Vf_nds > Vf_sup:
                i+=1
                Vf_inf = v_bateau[i]
                Vf_sup = v_bateau[i+1]
            # Vf_inf et Vf_sup encadrent la valeur de la vitesse du bateau, par des valeurs où la polaire du safran est connue.
            if alpha > np.deg2rad(19.9):
                alpha = np.deg2rad(19.9)
            Lignes1, Lignes2 = fctAnnexes.trouver_lignes(workbook, alpha, [Vf_inf, Vf_sup])
            sheet_inf = workbook[str(Vf_inf) + "kts"]  # On ouvre la feuille de calcul correspondant a Vfinf
            sheet_sup = workbook[str(Vf_sup) + "kts"]  # On ouvre la feuille de calcul correspondant a Vfsup
            n1, n2 = Lignes1[0], Lignes1[1] #Pour la feuille de calcul Vf_inf
            m1, m2 = Lignes2[0], Lignes2[1] #Pour la feuille de calcul Vf_sup
            # Extraction données du fichier excel
            alpha_inf_Vfinf = sheet_inf["B" + str(n1)].value
            alpha_sup_Vfinf = sheet_inf["B" + str(n2)].value
            alpha_inf_Vfsup = sheet_sup["B" + str(m1)].value
            alpha_sup_Vfsup = sheet_sup["B" + str(m2)].value
            CL_inf_Vfinf = sheet_inf["F" + str(n1)].value
            CL_sup_Vfinf = sheet_inf["F" + str(n2)].value
            CL_inf_Vfsup = sheet_sup["F" + str(m1)].value
            CL_sup_Vfsup = sheet_sup["F" + str(m2)].value
            CD_inf_Vfinf = sheet_inf["H" + str(n1)].value
            CD_sup_Vfinf = sheet_inf["H" + str(n2)].value
            CD_inf_Vfsup = sheet_sup["H" + str(m1)].value
            CD_sup_Vfsup = sheet_sup["H" + str(m2)].value
            XCP_inf_Vfinf = sheet_inf["J" + str(n1)].value
            XCP_sup_Vfinf = sheet_inf["J" + str(n2)].value
            XCP_inf_Vfsup = sheet_sup["J" + str(m1)].value
            XCP_sup_Vfsup = sheet_sup["J" + str(m2)].value
            ZCP_inf_Vfinf = sheet_inf["K" + str(n1)].value
            ZCP_sup_Vfinf = sheet_inf["K" + str(n2)].value
            ZCP_inf_Vfsup = sheet_sup["K" + str(m1)].value
            ZCP_sup_Vfsup = sheet_sup["K" + str(m2)].value

            #NB : la surface est prise que pour une des valeurs de vitesse du bateau.

            # Première interpolation selon les angles, sur chaque feuille de calcul séparément
            CL_Vfinf = fctAnnexes.interpol_lin(alpha, alpha_inf_Vfinf, alpha_sup_Vfinf, CL_inf_Vfinf, CL_sup_Vfinf)
            CL_Vfsup = fctAnnexes.interpol_lin(alpha, alpha_inf_Vfsup, alpha_sup_Vfsup, CL_inf_Vfsup, CL_sup_Vfsup)
            CD_Vfinf = fctAnnexes.interpol_lin(alpha, alpha_inf_Vfinf, alpha_sup_Vfinf, CD_inf_Vfinf, CD_sup_Vfinf)
            CD_Vfsup = fctAnnexes.interpol_lin(alpha, alpha_inf_Vfsup, alpha_sup_Vfsup, CD_inf_Vfsup, CD_sup_Vfsup)
            XCP_Vfinf = fctAnnexes.interpol_lin(alpha, alpha_inf_Vfinf, alpha_sup_Vfinf, XCP_inf_Vfinf, XCP_sup_Vfinf)
            XCP_Vfsup = fctAnnexes.interpol_lin(alpha, alpha_inf_Vfsup, alpha_sup_Vfsup, XCP_inf_Vfsup, XCP_sup_Vfsup)
            ZCP_Vfinf = fctAnnexes.interpol_lin(alpha, alpha_inf_Vfinf, alpha_sup_Vfinf, ZCP_inf_Vfinf, ZCP_sup_Vfinf)
            ZCP_Vfsup = fctAnnexes.interpol_lin(alpha, alpha_inf_Vfsup, alpha_sup_Vfsup, ZCP_inf_Vfsup, ZCP_sup_Vfsup)
            # Seconde interpolation, selon les vitesses
            CL = fctAnnexes.interpol_lin(Vf_nds, Vf_inf, Vf_sup, CL_Vfinf, CL_Vfsup)
            CD = fctAnnexes.interpol_lin(Vf_nds, Vf_inf, Vf_sup, CD_Vfinf, CD_Vfsup)
            XCP = fctAnnexes.interpol_lin(Vf_nds, Vf_inf, Vf_sup, XCP_Vfinf, XCP_Vfsup)
            ZCP = fctAnnexes.interpol_lin(Vf_nds, Vf_inf, Vf_sup, ZCP_Vfinf, ZCP_Vfsup)
            CP_repsafran = np.array([XCP, 0, ZCP]).reshape(-1, 1)

    S = 0.04
    L = 0.5 * rho_eau * CL * S * Vs ** 2
    D = 0.5 * rho_eau * CD * S * Vs ** 2
    torseur_hydro_repFluide = np.concatenate((np.array([-D, L, 0]).reshape(-1, 1), np.array([0, 0, 0]).reshape(-1, 1)),
                                           axis=1)
    return torseur_hydro_repFluide, CP_repsafran

def chgt_base_pt_application(CP_repsafran, O_repsafran_repBateau, Lambda, phi):
    """
    Fonction qui passe les forces hydro du repère lié au safran à celui lié à l'avance du bateau.
    --------------------
    ENTREES :
    CP_repsafran : coordonnées du point d'application des forces véliques dans le repère lié à la voile
                  np.array de taille (3,1) au format [X_repsafran, Y_repsafran, Z_repsafran].T
    O_repsafran_repBateau : Vecteur donnant les coordonnées de l'origine du repère lié à la voile, dans le repère lié au bateau.
                           vecteur du même format que CP_repVoile
    Lambda : angle de dérive, angle entre axe x longitudinal du bateau et axe X d'avance du bateau
    phi : angle de gîte, angle entre la verticale terrestre et l'axe Z du repère d'avance du bateau
    --------------------
    SORTIES :
    CP_repAvBateau  : coordonnées du point d'application des forces hydro sur la safran dans le repère d'avance du bateau.
                      vecteur du même format que CP_repsafran.
    """
    ###############################################################
    # ETAPE 1 : Passage du repère lié à la safran à celui lié au bateau.
    # Définition de la matrice de passage de repVoile à repBateau
    CP_repBateau = CP_repsafran + O_repsafran_repBateau #On fait simplement une translation, pour passer le centre de poussée dans le repère bateau
    ###############################################################
    # ETAPE 2 : Passage du repère lié au bateau à celui lié à l'avance du bateau
    # Définition de la matrice de passage de repBateau à repAvBateau
    M_repBateau_repAvBateau = np.zeros((3,3))
    cL, sL, cP, sP = np.cos(Lambda), np.sin(Lambda), np.cos(phi), np.sin(phi)
    M_repBateau_repAvBateau[0][0] = cL
    M_repBateau_repAvBateau[0][1] = -cP*sL
    M_repBateau_repAvBateau[0][2] = sP*sL
    M_repBateau_repAvBateau[1][0] = sL
    M_repBateau_repAvBateau[1][1] = cL*cP
    M_repBateau_repAvBateau[1][2] = -sP*cL
    M_repBateau_repAvBateau[2][1] = sP
    M_repBateau_repAvBateau[2][2] = cP
    # Calcul des nouvelles coordonnées de CP
    CP_repAvBateau = M_repBateau_repAvBateau @ CP_repBateau
    return CP_repAvBateau

def chgt_base_forces(Vs, Lambda, phi, torseur_hydro_repFluide):
    """
    Fonction qui passe les forces hydro du repère lié à la safran à celui lié à l'avance du bateau.
    --------------------
    ENTREES :
    Vs : vitesse du bateau
    Lambda : angle de dérive, angle entre axe x longitudinal du bateau et axe X d'avance du bateau
    phi : angle de gîte, angle entre la verticale terrestre et l'axe Z du repère d'avance du bateau
    torseur_hydro_repFluide : torseur aérodynamique donné dans le repère associé à l'écoulement du fluide.
                            Torseur de la forme : [[Fx, Fy, Fz],[Mx, My, Mz]]
    --------------------
    SORTIES :
    F_repAvBateau : forces hydro dans le repère lié à l'avance du bateau. De la forme np.array([Fx,Fy,Fz].T)
    """
    Lsafran, Dsafran = np.abs(torseur_hydro_repFluide[1][0]), np.abs(torseur_hydro_repFluide[0][0])
    Lambda_phi = np.arctan(np.tan(Lambda) * np.cos(phi))
    # FX = -Lsafran*np.sin(Lambda-Lambda_phi)-Dsafran*np.cos(Lambda-Lambda_phi)
    # FY = np.cos(phi)*(Lsafran*np.cos(Lambda-Lambda_phi)-Dsafran*np.sin(Lambda-Lambda_phi))
    # FZ = np.sin(phi)*(Lsafran*np.cos(Lambda-Lambda_phi)-Dsafran*np.sin(Lambda-Lambda_phi))
    FX = -Dsafran
    FY = np.cos(phi) * (Lsafran) * np.sign(Lambda)  # Si lambda positif, la force est positive, si lambda est négatif la force est négative
    FZ = np.sin(phi) * (Lsafran)
    F_repAvBateau = np.array([FX, FY, FZ]).reshape(-1,1)
    return F_repAvBateau

def chgt_base_moments(CP_repAvBateau, F_repAvBateau):
    """
    Fonction qui passe les moments associés aux forces véliques, du repère lié au vent à celui lié à l'avance du bateau.
    --------------------
    ENTREES :
    CP_repAvBateau : coordonnées du centre de poussée dans le repère d'avance du bateau, sous la forme d'un vecteur de taille 3.
    F_repAvBateau : Force vélique exprimée dans le repère d'avance du bateau, sous la forme d'un vecteur de taille 3.
    --------------------
    SORTIES :
    M_repAvBateau : Moment en l'origine du repère d'avance du bateau dû aux forces véliques, exprimé sous forme d'un vecteur de taille 3
    """
    M_repAvBateau = np.cross(CP_repAvBateau.T,F_repAvBateau.T)
    return M_repAvBateau.reshape(-1,1)

def torseur_Fhydro_safran(Vs, Lambda, phi, rho_eau, workbook):
    """
    Fonction qui renvoie le torseur des Forces véliques exprimé dans le repère d'avance du bateau, en son origine.
    --------------------
    ENTREES :
    fichier : chaîne de caractère donnant le nom du fichier avec les données d'entrée
    --------------------
    SORTIES :
    T : Torseur des forces véliques dans le repère d'avance du bateau, en son origine. Exprimé sous la forme :
        np.array([[FX, FY, FZ], [MX, MY, MZ]])
    """


    O_repsafran_repBateau = np.array([0.37387,0,-0.08757]).reshape(-1,1) #Repères voile et bateau ont la même origine
    torseur_hydro_repFluide, CP_repsafran = polaire_safran(Vs, rho_eau, Lambda, phi, workbook)
    CP_repAvBateau = chgt_base_pt_application(CP_repsafran, O_repsafran_repBateau, Lambda, phi)
    F_repAvBateau = chgt_base_forces(Vs, Lambda, phi, torseur_hydro_repFluide)
    M_repAvBateau = chgt_base_moments(CP_repAvBateau, F_repAvBateau)
    T = np.concatenate((F_repAvBateau, M_repAvBateau), axis = 1)
    return T

if __name__ == "__main__":
    import os
    path = os.getcwd()
    pathdir = os.path.dirname(path)
    workbook = load_workbook(filename=os.path.join(pathdir,"tests/data.xlsx"), data_only=True)
    sheet = workbook.active
    Vs = sheet["B3"].value
    Lambda = sheet["C3"].value
    phi = sheet["E3"].value
    rho_eau =sheet["H3"].value
    workbookSafran = load_workbook(filename=os.path.join(pathdir,"data/Polaires_Safran_data2.xlsm"), data_only=True)
    T = torseur_Fhydro_safran(Vs, Lambda, phi, rho_eau, workbookSafran)
    print(T)
