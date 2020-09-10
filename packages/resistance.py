import numpy as np
from openpyxl import load_workbook
from packages import fctAnnexes

def efforts_resistance(Vs, Lambda, workbook):
    """
    Fonction qui renvoie le torseur résultant des efforts hydrodynamiques, au point O et dans le repère lié à l'avance du bateau.
    Les valeurs d'entrée utilisées sont issues du calcul de résistance par la série de Delft.
    --------------------
    ENTREES :
    Vs : vitesse du bateau, en m/s
    Lambda : angle de dérive du bateau
    --------------------
    SORTIES :
    torseur__repAvBateau  : Torseur résultant des efforts hydrodynamique, au point O et dans le repère lié à l'avance du bateau.
                              Type : numpy.ndarray de taille (3,2)
    """
    Vs_nds = fctAnnexes.ms2nds(Vs)


    sheet = workbook["Devis de masse"]
    lignes = fctAnnexes.trouver_lignes_resistance(workbook, "Devis de masse", Vs_nds)
    n1, n2 = lignes[0], lignes[1]
    Vs_inf = sheet["AC" + str(n1)].value
    Vs_sup = sheet["AC" + str(n2)].value
    R_inf = sheet["AE" + str(n1)].value
    R_sup = sheet["AE"+str(n2)].value
    R = fctAnnexes.interpol_lin(Vs_nds, Vs_inf, Vs_sup, R_inf, R_sup)
    torseur_resistance_repAvBateau = np.concatenate((np.array([-R, 0, 0]).reshape(-1, 1), np.array([0, 0, 0]).reshape(-1, 1)),
                                           axis=1)
    return torseur_resistance_repAvBateau

def torseur_Fresistance(Vs, Lambda, workbook):
    """
    Fonction qui à partir des données d'entrée contenues dans fichier, renvoie le torseur dû aux efforts hydrodynamique
    dans le repère d'avance du bateau
    --------------------
    ENTREES :
    fichier : chaîne de caractère donnant le nom du fichier avec les données d'entrée
   --------------------
    SORTIES :
    torseur_resistance_repAvBateau : Torseur résultant des efforts hydrodynamiques, dans le repère lié à l'avance du bateau.
                             Type : numpy.ndarray de taille (3,2)
    """

    torseur_resistance_repAvBateau = efforts_resistance(Vs, Lambda, workbook)

    return torseur_resistance_repAvBateau

if __name__ == "__main__":
    import os
    path = os.getcwd()
    pathdir = os.path.dirname(path)
    workbook = load_workbook(filename=os.path.join(pathdir,"tests/data.xlsx"), data_only=True)
    sheet = workbook.active
    Vs = sheet["B3"].value
    Lambda = sheet["C3"].value
    workbookDVP = load_workbook(filename=os.path.join(pathdir,"data/Devis_masse.xlsx"), data_only=True)
    T = torseur_Fresistance(Vs, Lambda, workbookDVP)
    print(T)
