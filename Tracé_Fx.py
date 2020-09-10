import numpy as np
from openpyxl import load_workbook
from packages import aero, quille, resistance, safran, stab, fctAnnexes
import matplotlib.pyplot as plt
# l'idée là est d'importer toutes les données nécessaires et ne variants pas.
#
rho_air = 1.225
rho_eau = 1025
DELTA = 72
theta = 0.017453293

workbookAero = load_workbook(filename="data/Polaires_Voile_data4.xlsm", data_only=True)
workbookQuille = load_workbook(filename="data/Polaires_Quille_data2.xlsm", data_only=True)
workbookDVP = load_workbook(filename="data/Devis_masse.xlsx", data_only=True)
workbookSafran = load_workbook(filename="data/Polaires_Safran_data2.xlsm", data_only=True)
workbookStab = load_workbook(filename="data/Stab_data.xlsx", data_only=True)

def Tracé_en_Lambda():
    V = fctAnnexes.nds2ms(3.94)
    PHI = np.deg2rad(-1.77)
    Vt = fctAnnexes.nds2ms(20)
    angle_allure = np.deg2rad(120)
    L_lambda_deg = np.linspace(-10,10,100)
    L_lambda = np.deg2rad(L_lambda_deg)
    L_Fxtot = []
    L_Fxaero = []
    L_Fxsafran = []
    L_Fxquille = []
    L_Fxstab = []
    L_Fxresistance = []
    for LAMBDA in L_lambda:
        T_aero = aero.torseur_Faero(Vt, V, angle_allure, LAMBDA, PHI, rho_eau, rho_air, workbookAero)
        T_safran = safran.torseur_Fhydro_safran(V, LAMBDA, PHI, rho_eau, workbookSafran)
        T_stab = stab.torseur_Fstab(DELTA, LAMBDA, PHI, theta, workbookStab)
        T_quille = quille.torseur_Fhydro_quille(V, LAMBDA, PHI, rho_eau, workbookQuille)
        T_resistance = resistance.torseur_Fresistance(V, LAMBDA, workbookDVP)
        Fx_aero = T_aero[0][0]
        Fx_safran = T_safran[0][0]
        Fx_quille = T_quille[0][0]
        Fx_resistance = T_resistance[0][0]
        Fx_stab = T_stab[0][0]
        SommeFx = Fx_aero + Fx_quille + Fx_resistance + Fx_safran + Fx_stab
        L_Fxtot.append(SommeFx)
        L_Fxaero.append(Fx_aero)
        L_Fxsafran.append(Fx_safran)
        L_Fxquille.append(Fx_quille)
        L_Fxstab.append(Fx_stab)
        L_Fxresistance.append(Fx_resistance)

    plt.plot(L_lambda_deg, L_Fxtot, label = "Fx_total")
    plt.plot(L_lambda_deg, L_Fxaero, label = "Fx_aero")
    plt.plot(L_lambda_deg, L_Fxquille, label = "Fx_quille")
    plt.plot(L_lambda_deg, L_Fxsafran, label = "Fx_safran")
    plt.plot(L_lambda_deg, L_Fxstab, label = "Fx_stab")
    plt.plot(L_lambda_deg, L_Fxresistance, label = "Fx_resistance")
    plt.ylabel("Fx (N)")
    plt.xlabel("Lambda (deg)")
    plt.legend()
    plt.show()

def Tracé_en_PHI():
    V = 0.8
    LAMBDA = np.deg2rad(3)
    Vt = fctAnnexes.nds2ms(5)
    angle_allure = np.deg2rad(120)
    L_phi_deg = np.linspace(-110,110,50)
    L_phi = np.deg2rad(L_phi_deg)
    L_Fxtot = []
    L_Fxaero = []
    L_Fxsafran = []
    L_Fxquille = []
    L_Fxstab = []
    L_Fxresistance = []
    for PHI in L_phi:
        print(PHI)
        T_aero = aero.torseur_Faero(Vt, V, angle_allure, LAMBDA, PHI, rho_eau, rho_air, workbookAero)
        T_safran = safran.torseur_Fhydro_safran(V, LAMBDA, PHI, rho_eau, workbookSafran)
        T_stab = stab.torseur_Fstab(DELTA, LAMBDA, PHI, theta, workbookStab)
        T_quille = quille.torseur_Fhydro_quille(V, LAMBDA, PHI, rho_eau, workbookQuille)
        T_resistance = resistance.torseur_Fresistance(V, LAMBDA, workbookDVP)
        Fx_aero = T_aero[0][0]
        Fx_safran = T_safran[0][0]
        Fx_quille = T_quille[0][0]
        Fx_resistance = T_resistance[0][0]
        Fx_stab = T_stab[0][0]
        SommeFx = Fx_aero + Fx_quille + Fx_resistance + Fx_safran + Fx_stab
        L_Fxtot.append(SommeFx)
        L_Fxaero.append(Fx_aero)
        L_Fxsafran.append(Fx_safran)
        L_Fxquille.append(Fx_quille)
        L_Fxstab.append(Fx_stab)
        L_Fxresistance.append(Fx_resistance)

    plt.plot(L_phi_deg, L_Fxtot, label = "Fx_total")
    plt.plot(L_phi_deg, L_Fxaero, label = "Fx_aero")
    plt.plot(L_phi_deg, L_Fxquille, label = "Fx_quille")
    plt.plot(L_phi_deg, L_Fxsafran, label = "Fx_safran")
    plt.plot(L_phi_deg, L_Fxstab, label = "Fx_stab")
    plt.plot(L_phi_deg, L_Fxresistance, label = "Fx_resistance")
    plt.ylabel("Fx (N)")
    plt.xlabel("Phi (deg)")
    plt.legend()
    plt.show()


def Tracé_en_Vs():
    PHI = np.deg2rad(1)
    Vt = fctAnnexes.nds2ms(20)
    angle_allure = np.deg2rad(120)
    LAMBDA = np.deg2rad(1)
    L_Vs_nds = np.linspace(0,10,50)
    L_Vs = [fctAnnexes.nds2ms(x) for x in L_Vs_nds]
    L_Fxtot = []
    L_Fxaero = []
    L_Fxsafran = []
    L_Fxquille = []
    L_Fxstab = []
    L_Fxresistance = []
    for V in L_Vs:
        print(V)
        T_aero = aero.torseur_Faero(Vt, V, angle_allure, LAMBDA, PHI, rho_eau, rho_air)
        T_safran = safran.torseur_Fhydro_safran(V, LAMBDA, PHI, rho_eau)
        T_stab = stab.torseur_Fstab(DELTA, LAMBDA, PHI, theta)
        T_quille = quille.torseur_Fhydro_quille(V, LAMBDA, PHI, rho_eau)
        T_resistance = resistance.torseur_Fresistance(V, LAMBDA)
        Fx_aero = T_aero[0][0]
        Fx_safran = T_safran[0][0]
        Fx_quille = T_quille[0][0]
        Fx_resistance = T_resistance[0][0]
        Fx_stab = T_stab[0][0]
        SommeFx = Fx_aero + Fx_quille + Fx_resistance + Fx_safran + Fx_stab
        L_Fxtot.append(SommeFx)
        L_Fxaero.append(Fx_aero)
        L_Fxsafran.append(Fx_safran)
        L_Fxquille.append(Fx_quille)
        L_Fxstab.append(Fx_stab)
        L_Fxresistance.append(Fx_resistance)

    plt.plot(L_Vs_nds, L_Fxtot, label = "Fx_total")
    plt.plot(L_Vs_nds, L_Fxaero, label = "Fx_aero")
    plt.plot(L_Vs_nds, L_Fxquille, label = "Fx_quille")
    plt.plot(L_Vs_nds, L_Fxsafran, label = "Fx_safran")
    plt.plot(L_Vs_nds, L_Fxstab, label = "Fx_stab")
    plt.plot(L_Vs_nds, L_Fxresistance, label = "Fx_resistance")
    plt.ylabel("Fx (N)")
    plt.xlabel("Vs (kt)")
    plt.legend()
    plt.show()

if __name__=="__main__":
    Tracé_en_PHI()
    plt.show()
