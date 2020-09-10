import os, sys
import numpy as np
from openpyxl import load_workbook
currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)
sys.path.append(parentdir)
from packages import fctAnnexes
import unittest

path = os.getcwd()
pathdir = os.path.dirname(path)

class testFonctions_annexes(unittest.TestCase):
    def testInterpol_lin(self):
        # Test cas standard, A et B donnés dans le bon ordre
        xA, yA = 3, 12
        xB, yB = 5, 14
        x1, x2, x3 = 3, 4, 5
        self.assertEqual(fctAnnexes.interpol_lin(x1, xA, xB, yA, yB), 12.)
        self.assertEqual(fctAnnexes.interpol_lin(x3, xA, xB, yA, yB), 14.)
        self.assertEqual(fctAnnexes.interpol_lin(x2, xA, xB, yA, yB), 13.)
        # Test cas opposé, A et B donnés dans le désordre
        xB, yB = 3, 12
        xA, yA = 5, 14
        x1, x2, x3 = 3, 4, 5
        self.assertEqual(fctAnnexes.interpol_lin(x1, xA, xB, yA, yB), 12.)
        self.assertEqual(fctAnnexes.interpol_lin(x3, xA, xB, yA, yB), 14.)
        self.assertEqual(fctAnnexes.interpol_lin(x2, xA, xB, yA, yB), 13.)

    def testTrouver_lignes(self):
        workbookAero = load_workbook(filename=os.path.join(pathdir,"data/Polaires_Voile_data4.xlsm"), data_only=True)
        #Cas de test standard, pas de particularités
        Lignes1 = fctAnnexes.trouver_lignes(workbookAero, np.deg2rad(8.2), 15)
        self.assertEqual(Lignes1, [19,20])
        Lignes2, Lignes3 = fctAnnexes.trouver_lignes(workbookAero, np.deg2rad(8.2), [5,10])
        self.assertEqual(Lignes2, [20,21]) # Lignes associées à la feuille 5kts
        self.assertEqual(Lignes3, [20,21]) # Lignes associées à la feuille 10kts
        #Cas où les lignes associées à la valeur d'angle, ne sont pas les mêmes sur la feuille de calcul pour cause de pb de convergence dans xfoil
        Lignes4, Lignes5 = fctAnnexes.trouver_lignes(workbookAero, np.deg2rad(10.2), [10,15])
        self.assertEqual(Lignes4, [24, 25])  # Lignes associées à la feuille 10kts
        self.assertEqual(Lignes5, [22, 23])  # Lignes associées à la feuille 15kts

    def testTrouver_lignes_stab(self):
        workbookStab = load_workbook(filename=os.path.join(pathdir,"data/Stab_data.xlsx"), data_only=True)
        Ligne = fctAnnexes.trouver_lignes_stab(workbookStab, "Trans", 47)
        self.assertEqual(Ligne, [11,12])
        Ligne = fctAnnexes.trouver_lignes_stab(workbookStab, "Trans", 0)
        self.assertEqual(Ligne, [2,3])
        Ligne = fctAnnexes.trouver_lignes_stab(workbookStab, "Trans", -47)
        self.assertEqual(Ligne, [11, 12])

    def testTrouver_lignes_resistance(self):
        workbookDVP = load_workbook(filename=os.path.join(pathdir,"data/Devis_masse.xlsx"), data_only=True)
        Ligne = fctAnnexes.trouver_lignes_resistance(workbookDVP,"Devis de masse", 2.6)
        self.assertEqual(Ligne, [14,15])