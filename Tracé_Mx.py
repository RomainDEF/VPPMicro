import numpy as np
from openpyxl import load_workbook
from packages import aero, quille, resistance, safran, stab, fctAnnexes
import matplotlib.pyplot as plt
# l'idée là est d'importer toutes les données nécessaires et ne variants pas.
#
rho_air = 1.225
rho_eau = 1025
DELTA = 72
theta = 0.017453293

workbookAero = load_workbook(filename="data/Polaires_Voile_data4.xlsm", data_only=True)
workbookQuille = load_workbook(filename="data/Polaires_Quille_data2.xlsm", data_only=True)
workbookDVP = load_workbook(filename="data/Devis_masse.xlsx", data_only=True)
workbookSafran = load_workbook(filename="data/Polaires_Safran_data2.xlsm", data_only=True)
workbookStab = load_workbook(filename="data/Stab_data.xlsx", data_only=True)

def Tracé_en_Lambda():
    V = 0.8
    PHI = -6E-3
    Vt = fctAnnexes.nds2ms(5)
    angle_allure = 120
    L_lambda_deg = np.linspace(-10,10,100)
    L_lambda = np.deg2rad(L_lambda_deg)
    L_Mxtot = []
    L_Mxaero = []
    L_Mxsafran = []
    L_Mxquille = []
    L_Mxstab = []
    L_Mxresistance = []
    for LAMBDA in L_lambda:
        print(LAMBDA)
        T_aero = aero.torseur_Faero(Vt, V, angle_allure, LAMBDA, PHI, rho_eau, rho_air, workbookAero)
        T_safran = safran.torseur_Fhydro_safran(V, LAMBDA, PHI, rho_eau, workbookSafran)
        T_stab = stab.torseur_Fstab(DELTA, LAMBDA, PHI, theta, workbookStab)
        T_quille = quille.torseur_Fhydro_quille(V, LAMBDA, PHI, rho_eau, workbookQuille)
        T_resistance = resistance.torseur_Fresistance(V, LAMBDA, workbookDVP)
        Mx_aero = T_aero[0][1]
        Mx_safran = T_safran[0][1]
        Mx_quille = T_quille[0][1]
        Mx_resistance = T_resistance[0][1]
        Mx_stab = T_stab[0][1]
        SommeMx = Mx_aero + Mx_quille + Mx_resistance + Mx_safran + Mx_stab
        L_Mxtot.append(SommeMx)
        L_Mxaero.append(Mx_aero)
        L_Mxsafran.append(Mx_safran)
        L_Mxquille.append(Mx_quille)
        L_Mxstab.append(Mx_stab)
        L_Mxresistance.append(Mx_resistance)

    plt.plot(L_lambda_deg, L_Mxtot, label = "Mx_total")
    plt.plot(L_lambda_deg, L_Mxaero, label = "Mx_aero")
    plt.plot(L_lambda_deg, L_Mxquille, label = "Mx_quille")
    plt.plot(L_lambda_deg, L_Mxsafran, label = "Mx_safran")
    plt.plot(L_lambda_deg, L_Mxstab, label = "Mx_stab")
    plt.plot(L_lambda_deg, L_Mxresistance, label = "Mx_resistance")
    plt.xlabel("Lambda (deg)")
    plt.ylabel("Mx (N.m)")
    plt.legend()
    plt.show()

def Tracé_en_PHI():
    V = 0.8
    LAMBDA = np.deg2rad(3)
    Vt = fctAnnexes.nds2ms(5)
    angle_allure = 120
    L_phi_deg = np.linspace(-50,50,50)
    L_phi = np.deg2rad(L_phi_deg)
    L_Mxtot = []
    L_Mxaero = []
    L_Mxsafran = []
    L_Mxquille = []
    L_Mxstab = []
    L_Mxresistance = []
    for PHI in L_phi:
        print(PHI)
        T_aero = aero.torseur_Faero(Vt, V, angle_allure, LAMBDA, PHI, rho_eau, rho_air, workbookAero)
        T_safran = safran.torseur_Fhydro_safran(V, LAMBDA, PHI, rho_eau, workbookSafran)
        T_stab = stab.torseur_Fstab(DELTA, LAMBDA, PHI, theta, workbookStab)
        T_quille = quille.torseur_Fhydro_quille(V, LAMBDA, PHI, rho_eau, workbookQuille)
        T_resistance = resistance.torseur_Fresistance(V, LAMBDA, workbookDVP)
        Mx_aero = T_aero[0][1]
        Mx_safran = T_safran[0][1]
        Mx_quille = T_quille[0][1]
        Mx_resistance = T_resistance[0][1]
        Mx_stab = T_stab[0][1]
        SommeMx = Mx_aero + Mx_quille + Mx_resistance + Mx_safran + Mx_stab
        L_Mxtot.append(SommeMx)
        L_Mxaero.append(Mx_aero)
        L_Mxsafran.append(Mx_safran)
        L_Mxquille.append(Mx_quille)
        L_Mxstab.append(Mx_stab)
        L_Mxresistance.append(Mx_resistance)

    plt.plot(L_phi_deg, L_Mxtot, label = "Mx_total")
    plt.plot(L_phi_deg, L_Mxaero, label = "Mx_aero")
    plt.plot(L_phi_deg, L_Mxquille, label = "Mx_quille")
    plt.plot(L_phi_deg, L_Mxsafran, label = "Mx_safran")
    plt.plot(L_phi_deg, L_Mxstab, label = "Mx_stab")
    plt.plot(L_phi_deg, L_Mxresistance, label = "Mx_resistance")
    plt.ylabel("Mx (N.m)")
    plt.xlabel("Phi (deg)")
    plt.legend()
    plt.show()

if __name__=="__main__":
    Tracé_en_PHI()
    Tracé_en_Lambda()
    plt.show()