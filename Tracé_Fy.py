import numpy as np
from openpyxl import load_workbook
from packages import aero, quille, resistance, safran, stab, fctAnnexes
import matplotlib.pyplot as plt
# l'idée là est d'importer toutes les données nécessaires et ne variants pas.
#
rho_air = 1.225
rho_eau = 1025
DELTA = 72
theta = 0.017453293

workbookAero = load_workbook(filename="data/Polaires_Voile_data4.xlsm", data_only=True)
workbookQuille = load_workbook(filename="data/Polaires_Quille_data2.xlsm", data_only=True)
workbookDVP = load_workbook(filename="data/Devis_masse.xlsx", data_only=True)
workbookSafran = load_workbook(filename="data/Polaires_Safran_data2.xlsm", data_only=True)
workbookStab = load_workbook(filename="data/Stab_data.xlsx", data_only=True)

def Tracé_en_Lambda():
    V = 0.8
    PHI = -6E-3
    Vt = fctAnnexes.nds2ms(5)
    angle_allure = 120
    L_lambda_deg = np.linspace(-10,10,100)
    L_lambda = np.deg2rad(L_lambda_deg)
    L_Fytot = []
    L_Fyaero = []
    L_Fysafran = []
    L_Fyquille = []
    L_Fystab = []
    L_Fyresistance = []
    for LAMBDA in L_lambda:
        T_aero = aero.torseur_Faero(Vt, V, angle_allure, LAMBDA, PHI, rho_eau, rho_air, workbookAero)
        T_safran = safran.torseur_Fhydro_safran(V, LAMBDA, PHI, rho_eau, workbookSafran)
        T_stab = stab.torseur_Fstab(DELTA, LAMBDA, PHI, theta, workbookStab)
        T_quille = quille.torseur_Fhydro_quille(V, LAMBDA, PHI, rho_eau, workbookQuille)
        T_resistance = resistance.torseur_Fresistance(V, LAMBDA, workbookDVP)
        Fy_aero = T_aero[1][0]
        Fy_safran = T_safran[1][0]
        Fy_quille = T_quille[1][0]
        Fy_resistance = T_resistance[1][0]
        Fy_stab = T_stab[1][0]
        SommeFy = Fy_aero + Fy_quille + Fy_resistance + Fy_safran + Fy_stab
        L_Fytot.append(SommeFy)
        L_Fyaero.append(Fy_aero)
        L_Fysafran.append(Fy_safran)
        L_Fyquille.append(Fy_quille)
        L_Fystab.append(Fy_stab)
        L_Fyresistance.append(Fy_resistance)

    plt.plot(L_lambda_deg, L_Fytot, label = "Fy_total")
    plt.plot(L_lambda_deg, L_Fyaero, label = "Fy_aero")
    plt.plot(L_lambda_deg, L_Fyquille, label = "Fy_quille")
    plt.plot(L_lambda_deg, L_Fysafran, label = "Fy_safran")
    plt.plot(L_lambda_deg, L_Fystab, label = "Fy_stab")
    plt.plot(L_lambda_deg, L_Fyresistance, label = "Fy_resistance")
    plt.ylabel("Fy (N.m)")
    plt.xlabel("Lambda (deg)")
    plt.legend()
    plt.show()

def Tracé_en_PHI():
    V = 0.8
    LAMBDA = np.deg2rad(3)
    Vt = fctAnnexes.nds2ms(5)
    angle_allure = 120
    L_phi_deg = np.linspace(-50,50,50)
    L_phi = np.deg2rad(L_phi_deg)
    L_Fytot = []
    L_Fyaero = []
    L_Fysafran = []
    L_Fyquille = []
    L_Fystab = []
    L_Fyresistance = []
    for PHI in L_phi:
        print(PHI)
        T_aero = aero.torseur_Faero(Vt, V, angle_allure, LAMBDA, PHI, rho_eau, rho_air, workbookAero)
        T_safran = safran.torseur_Fhydro_safran(V, LAMBDA, PHI, rho_eau, workbookSafran)
        T_stab = stab.torseur_Fstab(DELTA, LAMBDA, PHI, theta, workbookStab)
        T_quille = quille.torseur_Fhydro_quille(V, LAMBDA, PHI, rho_eau, workbookQuille)
        T_resistance = resistance.torseur_Fresistance(V, LAMBDA, workbookDVP)
        Fy_aero = T_aero[1][0]
        Fy_safran = T_safran[1][0]
        Fy_quille = T_quille[1][0]
        Fy_resistance = T_resistance[1][0]
        Fy_stab = T_stab[1][0]
        SommeFy = Fy_aero + Fy_quille + Fy_resistance + Fy_safran + Fy_stab
        L_Fytot.append(SommeFy)
        L_Fyaero.append(Fy_aero)
        L_Fysafran.append(Fy_safran)
        L_Fyquille.append(Fy_quille)
        L_Fystab.append(Fy_stab)
        L_Fyresistance.append(Fy_resistance)

    plt.plot(L_phi_deg, L_Fytot, label = "Fy_total")
    plt.plot(L_phi_deg, L_Fyaero, label = "Fy_aero")
    plt.plot(L_phi_deg, L_Fyquille, label = "Fy_quille")
    plt.plot(L_phi_deg, L_Fysafran, label = "Fy_safran")
    plt.plot(L_phi_deg, L_Fystab, label = "Fy_stab")
    plt.plot(L_phi_deg, L_Fyresistance, label = "Fy_resistance")
    plt.ylabel("Fy (N.m)")
    plt.xlabel("Phi (deg)")
    plt.legend()
    plt.show()

if __name__=="__main__":
    Tracé_en_PHI()
    plt.show()
