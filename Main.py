import numpy as np
from openpyxl import load_workbook
import scipy.optimize as scpo
from packages import aero, quille, resistance, safran, stab, dataWriting, fctAnnexes
import os
import warnings
warnings.simplefilter("default") # Change the filter in this process
os.environ["PYTHONWARNINGS"] = "default" # Also affect subprocesses
import matplotlib.pyplot as plt
import time

#Supprime l'avertissement concernant l'impossibilité pour openpyxl de lire les DataValidation d'Excel
warnings.filterwarnings("ignore", message="Data Validation extension is ")
warnings.filterwarnings("error", message="The iteration is not making good progress")

# Paramètres et fichiers de données
rho_air = 1.225
rho_eau = 1025
DELTA = 72
theta = 0.017453293
workbookAero = load_workbook(filename="data/Polaires_Voile_data4.xlsm", data_only=True)
workbookQuille = load_workbook(filename="data/Polaires_Quille_data2.xlsm", data_only=True)
workbookDVP = load_workbook(filename="data/Devis_masse.xlsx", data_only=True)
workbookSafran = load_workbook(filename="data/Polaires_Safran_data2.xlsm", data_only=True)
workbookStab = load_workbook(filename="data/Stab_data.xlsx", data_only=True)

def PFS(X, Vt, angle_allure):
    V = float(X[0])
    PHI = float(X[1])
    LAMBDA = float(X[2])
    # print("VALEURS EN DEBUT DE BOUCLE")
    #if V > fctAnnexes.nds2ms(10): #Polaires de safran et de quille ne sont plus prises en compte
    #    V = fctAnnexes.nds2ms(10)
    #    warnings.warn("Speed was above 10 kts and was reset to 10 kts.")
    #if PHI > np.deg2rad(90): #On empeche une gite trop importante
    #    PHI = np.deg2rad(60)
    #    warnings.warn("Heeling angle was above 90 degrees and was reset to 60 degrees")
    #if PHI < np.deg2rad(-90): #On empeche une contre gite trop importante
    #    PHI = np.deg2rad(-60)
    #    warnings.warn("Heeling angle was below -90 degrees and was reset to -60 degrees")

    # print("V=", ms2nds(V), "noeuds", "              PHI=", np.rad2deg(PHI), "degrés", "              LAMBDA=",
    # np.rad2deg(LAMBDA), "degrés")

    T_aero = aero.torseur_Faero(Vt, V, angle_allure, LAMBDA, PHI, rho_eau, rho_air, workbookAero)
    T_safran = safran.torseur_Fhydro_safran(V, LAMBDA, PHI, rho_eau, workbookSafran)
    T_stab = stab.torseur_Fstab(DELTA,LAMBDA, PHI, theta, workbookStab)
    T_quille = quille.torseur_Fhydro_quille(V,LAMBDA, PHI, rho_eau, workbookQuille)
    T_resistance = resistance.torseur_Fresistance(V, LAMBDA, workbookDVP)
    # print("T_aero","\n",T_aero,"___________","\n","T_safran","\n",T_safran,"___________","\n","T_stab","\n",T_stab,
    # "___________","\n","T_quille","\n",T_quille,"___________","\n","T_resistance","\n",T_resistance)

    SommeTorseurs = np.array([0., 0., 0.]).reshape((-1,))
    Fx_aero = T_aero[0][0]
    Fx_safran = T_safran[0][0]
    Fx_quille = T_quille[0][0]
    Fx_resistance = T_resistance[0][0]
    Fx_stab = T_stab[0][0]
    SommeTorseurs[0] = Fx_stab + Fx_safran + Fx_quille + Fx_resistance + Fx_aero

    Fy_aero = T_aero[1][0]
    Fy_safran = T_safran[1][0]
    Fy_quille = T_quille[1][0]
    Fy_resistance = T_resistance[1][0]
    Fy_stab = T_stab[1][0]
    SommeTorseurs[2] = Fy_stab + Fy_aero + Fy_resistance + Fy_safran + Fy_quille

    Mx_aero = T_aero[0][1]
    Mx_safran = T_safran[0][1]
    Mx_quille = T_quille[0][1]
    Mx_resistance = T_resistance[0][1]
    Mx_stab = T_stab[0][1]
    SommeTorseurs[1] = Mx_aero + Mx_quille + Mx_resistance + Mx_safran + Mx_stab

    # print("PFS=", SommeTorseurs)
    # if V > fctAnnexes.nds2ms(10): #Polaires de safran et de quille ne sont plus prises en compte
    #    V = fctAnnexes.nds2ms(10)

    LISTEV.append(V)
    LISTEPHI.append(PHI)
    LISTELAMBDA.append(LAMBDA)
    LISTE_Fx.append(SommeTorseurs[0])
    LISTE_Fy.append(SommeTorseurs[2])
    LISTE_Mx.append(SommeTorseurs[1])
    return 10*SommeTorseurs

if __name__ =="__main__":
    print("Lancement du Velocity Prediction Program")
    print("Initialisation...")
    #warnings.filterwarnings('ignore', 'The iteration is not making good progress')

    # Paramètres de la série de simulations
    L_Vt = [5,10,15,20,25,30] #Liste des vitesses en noeud
    L_angles = [k for k in range(30, 181, 50)] #Liste des angles d'allure en degrés
    plot = True

    # Saisie du nom du répertoire d'enregistrement
    print("Saisir le nom ou le numéro de la série")
    nom_serie = input()
    t0 = time.time()

    # Création du répertoire d'enregistrement principal
    mainfilename = "Serie_" + nom_serie
    os.chdir('results')
    os.mkdir(mainfilename)

    # for Vt in L_Vt:
        # Création des répertoires d'enregistrement pour chaque vitesse
        # Vtfilename = "Vt_" + str(Vt) + "kt"
        # os.chdir(mainfilename)
        # os.mkdir(Vtfilename)
        # os.chdir("..")
        # for angle in L_angles:
        #     # Création des répertoires d'enregistrement pour chaque angle (dans chaque répertoir de vitesse)
        #     anglefilename = "AngleAllure_" + str(angle) + "deg"
        #     os.chdir(mainfilename)
        #     os.chdir(Vtfilename)
        #     os.mkdir(anglefilename)
        #     os.chdir("../..")

    print("Initialisation...OK \n")
    # Résolution de l'équilibre
    mainfilename = "Serie_" + nom_serie
    BILAN_BILAN_Vs = []
    BILAN_BILAN_Phi = []
    BILAN_BILAN_Lambda = []
    achieve=[]  # Permet de savoir quels angles ont été réussis
    fig = plt.figure()
    for Vt in L_Vt:
        print("Calculs pour Vt = "+str(Vt)+" kt")
        Vtfilename = "Vt_" + str(Vt) + "kt"
        # Paramètres pour initialiser la recherche de zéro
        Vs = 2
        Lambda = np.deg2rad(1)
        phi = np.deg2rad(20)
        X0 = np.array([Vs, phi, Lambda]).reshape((-1, 1))
        # Variables de stockage des résulats par valeur de vent
        BILAN_Vs = []
        BILAN_Phi = []
        BILAN_Lambda = []
        BILAN_Fx = []
        BILAN_Fy = []
        BILAN_Mx = []
        achieve.append([])

        for angle in L_angles:
            print("    Calculs en cours pour un angle d'allure de " + str(angle) + " deg")
            anglefilename = "AngleAllure_" + str(angle) + "deg"
            # Listes de stockage des variables au fur et à mesure de la convergence
            Vt_ms = fctAnnexes.nds2ms(Vt)
            angle_rad = np.deg2rad(angle)
            LISTEV = []
            LISTEPHI = []
            LISTELAMBDA = []
            LISTE_Fx =[]
            LISTE_Fy =[]
            LISTE_Mx =[]
            try:
                X_sol = scpo.fsolve(PFS, X0, args= (Vt_ms,angle_rad), xtol=1e-06, maxfev=500)
                achieve[-1].append(True)
            except(RuntimeWarning):
                try:
                    Vs = 2
                    Lambda = np.deg2rad(1)
                    phi = np.deg2rad(20)
                    X0 = np.array([Vs, phi, Lambda]).reshape((-1, 1))
                    X_sol = scpo.fsolve(PFS, X0, args=(Vt_ms, angle_rad), xtol=1e-06, maxfev=500)
                    achieve[-1].append(True)
                except:
                    print("Le process ne trouve pas le résultat, l'angle d'allure de " + str(angle) + " deg à Vt = " + str(Vt) + " kt n'est pas pris en compte pour les résultats")
                    achieve[-1].append(False)
                    continue
            BILAN_Vs.append(X_sol[0])
            BILAN_Phi.append(X_sol[1])
            BILAN_Lambda.append(X_sol[2])
            BILAN_Fx.append(LISTE_Fx[-1])
            BILAN_Fy.append(LISTE_Fy[-1])
            BILAN_Mx.append(LISTE_Mx[-1])
            X0 = np.array([BILAN_Vs[-1], BILAN_Phi[-1], BILAN_Lambda[-1]]).reshape((-1, 1))

            # path = mainfilename+"/"+Vtfilename+"/"+anglefilename
            # dataWriting.writeFile_rawdata("RESULTS",path,Vt,angle,[LISTEV,LISTEPHI,LISTELAMBDA,LISTE_Fx, LISTE_Fy, LISTE_Mx])

            # if plot:
            #     fig.suptitle("Vt="+str(Vt)+"kts - angle allure = "+str(angle)+" degrés")
            #     ax1 = fig.add_subplot(221)
            #     ax1.set(xlabel='Nombre d itérations', ylabel='Vitesse (m/s)')
            #     ax1.plot(LISTEV)
            #     ax2 = fig.add_subplot(222)
            #     ax2.set(xlabel='Nombre d itérations', ylabel='Gîte (rad)')
            #     ax2.plot(LISTEPHI)
            #     ax3 = fig.add_subplot(223)
            #     ax3.set(xlabel='Nombre d itérations''Dérive (rad)', ylabel='Dérive (rad)')
            #     ax3.plot(LISTELAMBDA)
            #     ax4 = fig.add_subplot(224)
            #     ax4.plot(LISTE_Fx, label ="Fx (N)")
            #     ax4.plot(LISTE_Fy, label ="Fy (N)")
            #     ax4.plot(LISTE_Mx, label ="Mx (N.m)")
            #     ax4.legend()
            #     ax4.set(xlabel='Nombre d itérations', ylabel='Forces et moment')
            #     ax1.title.set_text('VITESSE')
            #     ax2.title.set_text('GITE')
            #     ax3.title.set_text('DERIVE')
            #     ax4.title.set_text('EFFORTS/MOMENTS SUR LE BATEAU')
                ##plt.show()
                # os.chdir(path)
                # plt.savefig("Convergence_"+str(Vt)+"kt_"+str(angle)+"deg.png", format="png")
                # plt.clf()
                # os.chdir("../../..")
        #     print("        Calculs pour un angle d'allure de " + str(angle) + " deg...OK")
        # print("Calculs pour Vt = "+str(Vt)+" kt...OK")
        # print("Enregistrement des fichiers de sortie pour cette vitesse de vent")
        BILAN_BILAN_Vs.append(BILAN_Vs)
        BILAN_BILAN_Phi.append(BILAN_Phi)
        BILAN_BILAN_Lambda.append(BILAN_Lambda)

        # path = mainfilename + "/" + Vtfilename
        # dataWriting.writeFile_bilan("BILAN",path,Vt,[L_angles, BILAN_Vs, BILAN_Phi, BILAN_Lambda, BILAN_Fx, BILAN_Fy, BILAN_Mx])
        #
        # if plot:
        #     L_theta = list(np.deg2rad(L_angles.copy()))
        #     L_Vs = BILAN_Vs.copy()
        #     p = len(L_theta)
        #     for i in range(p-1,-1,-1):
        #         L_theta.append(-L_theta[i])
        #         L_Vs.append(L_Vs[i])
        #
        #     ax = plt.subplot(111,polar=True)
        #     ax.plot(L_theta, L_Vs)
        #     ax.set_theta_zero_location("N")
        #     ax.set_theta_direction('clockwise')
        #
        #     os.chdir(path)
        #     plt.savefig("Polaires_"+str(Vt)+"kt.png", format="png")
        #     plt.clf()
        #     os.chdir("../..")
        # print("Enregistrement des fichiers de sortie pour cette vitesse de vent...OK")
        # print()

    if plot:
        L_theta = list(np.deg2rad(L_angles.copy()))
        p = len(L_theta)
        p_achieve = []
        for a in BILAN_BILAN_Vs:
            p_achieve.append(len(a))
        for i in range(p-1, -1, -1):
            L_theta.append(-L_theta[i])
            for L in achieve:
                L.append(L[i])

        ax = plt.subplot(111, polar=True)
        ax.set_theta_zero_location("N")
        ax.set_theta_direction('clockwise')

        for k in range(len(BILAN_BILAN_Vs)):
            L_Vs = BILAN_BILAN_Vs[k].copy()
            L_theta_True = [i for (i, v) in zip(L_theta, achieve[k]) if v] # Liste des angles réussis
            for i in range(p_achieve[k]-1, -1, -1):
                L_Vs.append(L_Vs[i])
            ax.plot(L_theta_True, L_Vs, label = str(L_Vt[k])+" kt")
        plt.legend()

        os.chdir(mainfilename)
        plt.savefig("Polaires_voilier.png", format="png", dpi=600)
        plt.clf()

        ax = plt.subplot(111)
        ax.grid(True)
        for k in range(len(BILAN_BILAN_Phi)):
            L_Vs = np.rad2deg(BILAN_BILAN_Phi[k].copy())
            L_angles_True = [i for (i, v) in zip(L_angles, achieve[k]) if v]
            ax.plot(L_angles_True, L_Vs, label=str(L_Vt[k]) + " kt")
            ax.set_title("Angle de gîte en fonction de l'angle d'allure")
            ax.set_xlabel("Angle d'allure [deg]")
            ax.set_ylabel("Angle de gîte [deg]")
        plt.legend()

        plt.savefig("Phi_voilier.png", format="png", dpi=600)
        plt.clf()

        ax = plt.subplot(111)
        ax.grid(True)
        for k in range(len(BILAN_BILAN_Lambda)):
            L_Vs = np.rad2deg(BILAN_BILAN_Lambda[k].copy())
            L_angles_True = [i for (i, v) in zip(L_angles, achieve[k]) if v]
            ax.plot(L_angles_True, L_Vs, label=str(L_Vt[k]) + " kt")
            ax.set_title("Angle de dérive en fonction de l'angle d'allure")
            ax.set_xlabel("Angle d'allure [deg]")
            ax.set_ylabel("Angle de dérive [deg]")
        plt.legend()

        plt.savefig("Lambda_voilier.png", format="png", dpi=600)
        plt.clf()
    tfin = time.time()
    print("Calcul terminé en "+str(round(tfin-t0,1))+" secondes")

